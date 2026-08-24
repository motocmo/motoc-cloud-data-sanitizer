from __future__ import annotations

import csv
import io
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from cloud_data_sanitizer.detection import detect_field
from cloud_data_sanitizer.models import Dataset, SanitizerError, SheetInfo

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
MAX_INPUT_BYTES = 100 * 1024 * 1024
MAX_WORKBOOK_UNCOMPRESSED_BYTES = 500 * 1024 * 1024
MAX_WORKBOOK_ENTRIES = 20_000


def _validate_path(path: Path) -> str:
    if not path.is_file():
        raise SanitizerError("input_missing", "Input file does not exist.")
    extension = path.suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise SanitizerError(
            "unsupported_type",
            "Supported file types are .csv and .xlsx.",
        )
    if path.stat().st_size == 0:
        raise SanitizerError("input_empty", "The input file is empty.")
    if path.stat().st_size > MAX_INPUT_BYTES:
        raise SanitizerError(
            "input_too_large",
            "The input file exceeds the 100 MB local processing limit.",
        )
    return extension


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SanitizerError("csv_encoding", "The CSV encoding could not be detected.")


def _read_csv(path: Path) -> Dataset:
    text = _decode_csv(path.read_bytes())
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise SanitizerError("csv_no_header", "The CSV does not contain a header row.")
    headers = [str(header).strip() for header in reader.fieldnames]
    rows = [
        {
            header: row.get(original)
            for header, original in zip(headers, reader.fieldnames, strict=True)
        }
        for row in reader
        if any(value is not None and str(value).strip() for value in row.values())
    ]
    if not rows:
        raise SanitizerError(
            "csv_no_rows",
            "The CSV contains headers but no data rows.",
        )
    return Dataset(path, "csv", headers, rows, delimiter=dialect.delimiter)


def _open_workbook(path: Path) -> Any:
    try:
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            total = sum(entry.file_size for entry in entries)
            if (
                len(entries) > MAX_WORKBOOK_ENTRIES
                or total > MAX_WORKBOOK_UNCOMPRESSED_BYTES
            ):
                raise SanitizerError(
                    "xlsx_expansion",
                    "The XLSX workbook expands beyond the safe processing limit.",
                )
        # data_only=True reads cached values; formulas are never evaluated.
        return load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except SanitizerError:
        raise
    except (
        zipfile.BadZipFile,
        InvalidFileException,
        OSError,
        ValueError,
        KeyError,
    ) as exc:
        raise SanitizerError(
            "xlsx_invalid",
            "The XLSX workbook is invalid or could not be opened.",
        ) from exc


def _worksheet_rows(worksheet: Any) -> tuple[list[str], list[dict[str, Any]]]:
    nonempty = [
        row
        for row in worksheet.iter_rows(values_only=True)
        if any(value is not None and str(value).strip() for value in row)
    ]
    if not nonempty:
        return [], []
    header_index = 0
    best_score = -1
    for index, row in enumerate(nonempty[:20]):
        score = sum(
            detect_field(str(value or ""), fuzzy=False) is not None for value in row
        )
        if score > best_score:
            header_index, best_score = index, score
    raw_headers = [
        str(value).strip() if value is not None else ""
        for value in nonempty[header_index]
    ]
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, header in enumerate(raw_headers):
        base = header or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    rows: list[dict[str, Any]] = []
    for raw in nonempty[header_index + 1 :]:
        padded = tuple(raw) + (None,) * max(0, len(headers) - len(raw))
        rows.append(dict(zip(headers, padded, strict=False)))
    return headers, rows


def inspect_sheets(path: str | Path) -> list[SheetInfo]:
    source = Path(path).expanduser().resolve()
    extension = _validate_path(source)
    if extension == ".csv":
        dataset = _read_csv(source)
        fields = tuple(
            sorted(filter(None, (detect_field(header) for header in dataset.headers)))
        )
        return [
            SheetInfo(
                "CSV data",
                len(dataset.rows),
                len(dataset.headers),
                fields,
                "cost" in fields,
            )
        ]
    workbook = _open_workbook(source)
    try:
        result: list[SheetInfo] = []
        for worksheet in workbook.worksheets:
            headers, rows = _worksheet_rows(worksheet)
            fields = tuple(
                sorted(filter(None, (detect_field(header) for header in headers)))
            )
            result.append(
                SheetInfo(
                    worksheet.title,
                    len(rows),
                    len(headers),
                    fields,
                    bool(rows) and "cost" in fields,
                )
            )
        return result
    finally:
        workbook.close()


def read_dataset(path: str | Path, sheet_name: str | None = None) -> Dataset:
    source = Path(path).expanduser().resolve()
    extension = _validate_path(source)
    if extension == ".csv":
        if sheet_name:
            raise SanitizerError(
                "sheet_csv",
                "--sheet is only valid for XLSX workbooks.",
            )
        return _read_csv(source)
    workbook = _open_workbook(source)
    try:
        nonempty: list[tuple[str, list[str], list[dict[str, Any]]]] = []
        for worksheet in workbook.worksheets:
            headers, rows = _worksheet_rows(worksheet)
            if rows:
                nonempty.append((worksheet.title, headers, rows))
        if sheet_name:
            selected = next((item for item in nonempty if item[0] == sheet_name), None)
            if not selected:
                raise SanitizerError(
                    "sheet_missing",
                    "Selected worksheet does not exist or has no data rows.",
                    sheet=sheet_name,
                )
        elif len(nonempty) == 1:
            selected = nonempty[0]
        elif not nonempty:
            raise SanitizerError(
                "xlsx_empty",
                "The XLSX workbook contains no data rows.",
            )
        else:
            names = ", ".join(item[0] for item in nonempty)
            raise SanitizerError(
                "sheet_required",
                f"Multiple data worksheets detected ({names}). Select one with --sheet.",
            )
        name, headers, rows = selected
        return Dataset(source, "xlsx", headers, rows, sheet_name=name)
    finally:
        workbook.close()
