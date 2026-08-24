from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from cloud_data_sanitizer.checksum import sha256_file, verify_sha256_sidecar
from cloud_data_sanitizer.models import SanitizerError
from cloud_data_sanitizer.readers import read_dataset
from cloud_data_sanitizer.service import inspect_dataset, sanitize_dataset

KNOWN_KEY = bytes.fromhex("22" * 32)


def write_csv(path: Path, *, with_secret: bool = False) -> None:
    fieldnames = [
        "Service",
        "Cost",
        "Currency",
        "UsageDate",
        "ResourceId",
        "Resource Group",
    ]
    rows = [
        {
            "Service": "Compute",
            "Cost": "100.25",
            "Currency": "USD",
            "UsageDate": "2026-07-01",
            "ResourceId": "vm-prod-01",
            "Resource Group": "payments-prod",
        },
        {
            "Service": "Storage",
            "Cost": "-5.25",
            "Currency": "USD",
            "UsageDate": "2026-07-31",
            "ResourceId": "disk-prod-01",
            "Resource Group": "payments-prod",
        },
    ]
    if with_secret:
        fieldnames.append("AccessKey")
        for row in rows:
            row["AccessKey"] = "AKIAEXAMPLEKEY"
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def test_end_to_end_csv_pseudonymization_and_sha256(tmp_path: Path) -> None:
    source = tmp_path / "billing.csv"
    output = tmp_path / "billing_sanitized.csv"
    write_csv(source)
    original_bytes = source.read_bytes()

    result = sanitize_dataset(
        source,
        output,
        rules={"Resource Group": "pseudonymize"},
        customer_key=KNOWN_KEY,
        provider="azure",
    )

    assert source.read_bytes() == original_bytes
    original = read_dataset(source)
    sanitized = read_dataset(output)
    assert sanitized.rows[0]["Cost"] == original.rows[0]["Cost"]
    assert sanitized.rows[0]["ResourceId"].startswith("resource-id-")
    assert sanitized.rows[0]["ResourceId"] != original.rows[0]["ResourceId"]
    assert all(check.passed for check in result.validation)
    assert verify_sha256_sidecar(output)
    assert result.sha256 == sha256_file(output)
    report = json.loads(result.report_path.read_text(encoding="utf-8"))
    assert report["processing"] == "local_only"
    assert "HMAC-SHA-256" in report["pseudonymization"]
    assert report["mapping_generated"] is False


def test_xlsx_roundtrip(tmp_path: Path) -> None:
    source = tmp_path / "billing.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Azure"
    sheet.append(["Service", "Cost", "Currency", "UsageDate", "ResourceId"])
    sheet.append(["Compute", "10.50", "USD", "2026-07-01", "azure-vm"])
    workbook.save(source)
    workbook.close()

    output = tmp_path / "out.xlsx"
    result = sanitize_dataset(
        source,
        output,
        keep_potential=True,
        customer_key=KNOWN_KEY,
        provider="azure",
    )
    sanitized = read_dataset(output)
    assert sanitized.rows[0]["Cost"] in {"10.50", 10.5}
    assert sanitized.rows[0]["ResourceId"].startswith("resource-id-")
    assert verify_sha256_sidecar(result.output_path)


def test_restricted_column_is_physically_dropped(tmp_path: Path) -> None:
    source = tmp_path / "billing.csv"
    output = tmp_path / "out.csv"
    write_csv(source, with_secret=True)
    result = sanitize_dataset(
        source,
        output,
        rules={"Resource Group": "keep"},
        allow_remove=True,
        customer_key=KNOWN_KEY,
    )
    sanitized = read_dataset(output)
    assert "AccessKey" not in sanitized.headers
    assert "AccessKey" in result.dropped_columns
    text = output.read_text(encoding="utf-8")
    assert "AKIAEXAMPLEKEY" not in text


def test_potential_columns_require_explicit_decision(tmp_path: Path) -> None:
    source = tmp_path / "billing.csv"
    write_csv(source)
    with pytest.raises(SanitizerError, match="require a decision"):
        sanitize_dataset(source, tmp_path / "out.csv", customer_key=KNOWN_KEY)


def test_remove_requires_confirmation(tmp_path: Path) -> None:
    source = tmp_path / "billing.csv"
    write_csv(source)
    with pytest.raises(SanitizerError, match="allow-remove"):
        sanitize_dataset(
            source,
            tmp_path / "out.csv",
            rules={"Resource Group": "keep", "ResourceId": "remove"},
            customer_key=KNOWN_KEY,
        )


def test_original_path_cannot_be_overwritten(tmp_path: Path) -> None:
    source = tmp_path / "billing.csv"
    write_csv(source)
    original = source.read_bytes()
    with pytest.raises(SanitizerError, match="never modified"):
        sanitize_dataset(source, source, keep_potential=True, customer_key=KNOWN_KEY)
    assert source.read_bytes() == original


def test_multiple_excel_sheets_require_selection(tmp_path: Path) -> None:
    source = tmp_path / "billing.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Azure"
    first.append(["Service", "Cost", "ResourceId"])
    first.append(["Compute", 10, "azure-vm"])
    second = workbook.create_sheet("AWS")
    second.append(["Service", "Cost", "AccountId"])
    second.append(["EC2", 20, "123"])
    workbook.save(source)
    workbook.close()

    inspection = inspect_dataset(source)
    assert inspection["warning"] is not None
    with pytest.raises(SanitizerError, match="Select one with --sheet"):
        read_dataset(source)
    selected = read_dataset(source, "AWS")
    assert selected.rows[0]["Service"] == "EC2"


def test_sha256_detects_tamper(tmp_path: Path) -> None:
    source = tmp_path / "billing.csv"
    output = tmp_path / "out.csv"
    write_csv(source)
    sanitize_dataset(
        source,
        output,
        keep_potential=True,
        customer_key=KNOWN_KEY,
    )
    assert verify_sha256_sidecar(output)
    output.write_text(output.read_text(encoding="utf-8") + "\n", encoding="utf-8")
    assert verify_sha256_sidecar(output) is False
